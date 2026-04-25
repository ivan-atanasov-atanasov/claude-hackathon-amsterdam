"""
Populate buurt_baseline with Veiligheidsindex 2025-3 scores.

If the XLSX is provided (downloaded from https://onderzoek.amsterdam.nl/dataset/openbare-orde-en-veiligheid),
it is parsed and scores are loaded per buurt_code. Otherwise, all buurten from
buurt_polygons.json are inserted with a neutral default score of 0.7.

Usage:
    cd backend && source venv/bin/activate
    python -m data_import.import_veiligheidsindex [--xlsx path/to/file.xlsx] [--dry-run]
"""

import argparse
import json
import os
from pathlib import Path

from dotenv import load_dotenv
from supabase import create_client

load_dotenv()

DATA_DIR = Path(__file__).parent.parent / "data"
DEFAULT_SCORE = 0.7
CHUNK_SIZE = 200


def load_from_xlsx(path: str) -> dict[str, float]:
    """Parse Veiligheidsindex XLSX and return {buurt_code: normalized_score}."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Install openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    scores: dict[str, float] = {}
    header = [str(c.value).lower().strip() if c.value else "" for c in next(ws.iter_rows())]

    code_col = next((i for i, h in enumerate(header) if "code" in h or "buurt" in h), None)
    score_col = next((i for i, h in enumerate(header) if "index" in h or "score" in h or "veilig" in h), None)

    if code_col is None or score_col is None:
        raise ValueError(f"Cannot locate buurt_code/score columns. Headers: {header}")

    raw_scores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[code_col]).strip() if row[code_col] else None
        val = row[score_col]
        if code and val is not None:
            try:
                raw_scores.append((code, float(val)))
            except (ValueError, TypeError):
                pass

    if not raw_scores:
        return {}

    # Normalize to 0–1
    values = [s for _, s in raw_scores]
    min_v, max_v = min(values), max(values)
    rng = max_v - min_v or 1.0
    scores = {code: round((v - min_v) / rng, 4) for code, v in raw_scores}
    return scores


def load_defaults_from_buurt_polygons() -> list[str]:
    """Return all buurt_codes from the local buurt_polygons.json."""
    path = DATA_DIR / "buurt_polygons.json"
    if not path.exists():
        raise FileNotFoundError("Run import_gebieden.py first to generate data/buurt_polygons.json")
    with open(path) as f:
        data = json.load(f)
    return [feat["buurt_code"] for feat in data if feat.get("buurt_code")]


def build_records(scores: dict[str, float] | None, buurt_codes: list[str]) -> list[dict]:
    records = []
    for code in buurt_codes:
        records.append({
            "buurt_code": code,
            "veiligheidsindex": scores.get(code, DEFAULT_SCORE) if scores else DEFAULT_SCORE,
        })
    return records


def upsert(records: list[dict], dry_run: bool) -> None:
    if dry_run:
        print(f"[dry-run] Would upsert {len(records)} buurt_baseline rows.")
        if records:
            print("  Sample:", records[:3])
        return

    client = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"])
    total = 0
    for i in range(0, len(records), CHUNK_SIZE):
        chunk = records[i : i + CHUNK_SIZE]
        client.table("buurt_baseline").upsert(chunk, on_conflict="buurt_code").execute()
        total += len(chunk)
        print(f"  Upserted {total}/{len(records)}...")
    print(f"Done. {total} buurt_baseline rows loaded.")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", help="Path to Veiligheidsindex XLSX file (optional).")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    scores: dict[str, float] | None = None
    if args.xlsx:
        print(f"Parsing XLSX: {args.xlsx}")
        scores = load_from_xlsx(args.xlsx)
        print(f"  Loaded {len(scores)} buurt scores from XLSX.")
    else:
        print("No XLSX provided — using default score (0.7) for all buurten.")

    print("Loading buurt codes from buurt_polygons.json...")
    buurt_codes = load_defaults_from_buurt_polygons()
    print(f"  {len(buurt_codes)} buurten found.")

    records = build_records(scores, buurt_codes)
    upsert(records, args.dry_run)


if __name__ == "__main__":
    main()
